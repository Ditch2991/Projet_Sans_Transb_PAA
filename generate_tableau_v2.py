"""
generate_tableau_v2.py — Exports Excel App V2
==============================================
Structure NT + Transbordé séparés pour marchandises
Conteneurs : NT pilote × ratio
Escales    : identique App V1

Exports :
  generate_xlsx_lt_v2()  → Long terme  (2026-2040)
  generate_xlsx_ct_v2()  → Court terme (mensuel 2026)
"""

import io, pickle, numpy as np
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              numbers as xl_num)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────
# 0. COULEURS & STYLES — Palette V2 (inspirée V1, distincts)
# ─────────────────────────────────────────────────────────────────
# En-tête & titre
C_TOTAL   = "1B2631"   # Presque noir — en-têtes et titres sections
C_TOTAL_FG= "FFFFFF"
# Lignes Total (ligne principale avec réalisé)
C_TOTAL_LINE = "1ABC9C"  # Teal — ligne Total principale
C_TOTAL_LINE_FG = "000000"
C_REEL_FG = "FF0000"     # Rouge vif pour réalisés (comme V1)
# Marchandises
C_NT      = "154360"   # Bleu marine — Non transbordé
C_NT_FG   = "FFFFFF"
C_TR      = "BA4A00"   # Brun-rouge — Transbordé
C_TR_FG   = "FFFFFF"
C_SENS    = "1A5276"   # Bleu foncé — groupes sens/composante
C_SENS_FG = "FFFFFF"
C_COMP    = "6C3483"   # Violet — composante (≠ V1 violet A02B93)
C_COMP_FG = "FFFFFF"
C_CONT_C  = "117A65"   # Vert — conteneurisation
C_CONT_FG = "FFFFFF"
# Escales
C_ESC     = "0E6655"   # Vert foncé — escales (différent de V1 bleu)
C_ESC_FG  = "FFFFFF"
# Conteneurs
C_CNT_TOT = "1B2631"   # Presque noir
C_CNT_FG  = "FFFFFF"
C_CNT_NT  = "154360"   # Bleu marine — NT conteneurs
C_CNT_TR  = "784212"   # Brun — transbordé conteneurs
C_CNT_TR_FG="FFFFFF"
# Détails (lignes blanches)
C_BLANC   = "FFFFFF"
C_PREV_FG = "1B2631"
# Alias compatibilité
C_GRP1    = "154360"
C_GRP1_FG = "FFFFFF"
C_GRP2    = "BA4A00"
C_GRP2_FG = "FFFFFF"

def _fill(ws, row, col, color):
    ws.cell(row, col).fill = PatternFill("solid", fgColor=color)

def _c(ws, row, col, value="", bg=C_BLANC, fg="000000",
       bold=False, italic=False, size=10, align="center",
       num_fmt=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill  = PatternFill("solid", fgColor=bg)
    cell.font  = Font(color=fg, bold=bold, italic=italic, size=size,
                      name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center",
                                wrap_text=True)
    if num_fmt:
        cell.number_format = num_fmt
    if border:
        side = Side(style="thin", color="D0D0D0")
        cell.border = Border(left=side, right=side, top=side, bottom=side)
    return cell

def _col_letter(n): return get_column_letter(n)

# ─────────────────────────────────────────────────────────────────
# 1. CHARGEMENT DONNÉES
# ─────────────────────────────────────────────────────────────────
def _load_v2():
    with open("forecasts_v2.pkl","rb") as f:
        d = pickle.load(f)
    return d["forecasts"], d["meta"]

def _load_esc():
    with open("forecasts_escales.pkl","rb") as f:
        d = pickle.load(f)
    return d["forecasts"], d["meta"]

def _load_cnt_v2():
    with open("forecasts_conteneurs_v2.pkl","rb") as f:
        d = pickle.load(f)
    return d["forecasts"], d["meta"]

def _load_series_v2():
    with open("series_v2.pkl","rb") as f:
        return pickle.load(f)

# ─────────────────────────────────────────────────────────────────
# 2. HELPERS VALEURS
# ─────────────────────────────────────────────────────────────────
SEGS_NT_LABEL = {
    "Import"           : "Importations",
    "Export"           : "Exportations",
    "March. generales" : "March. générales",
    "Prod. petroliers" : "Prod. pétroliers",
    "Prod. de peche"   : "Prod. de pêche",
    "Conteneurise"     : "Conteneurisé",
    "Non conteneurise" : "Non conteneurisé",
}

def _nt_ann(fc_v2, series_v2, seg, yr, annee_fin):
    """Valeur annuelle NT ou segment NT."""
    if yr <= annee_fin:
        s = series_v2.get(seg)
        if s is None: return 0.0
        return float(s[s.index.year == yr].sum())
    fc = fc_v2.get(yr, {})
    if seg == "NT":
        return fc.get("annuel_nt", 0.0)
    if seg == "Transbordement":
        return fc.get("annuel_tr", 0.0)
    if seg == "Total":
        return fc.get("annuel_tot", 0.0)
    segs = fc.get("segs_nt", {})
    v = segs.get(seg)
    if v is None: return 0.0
    return float(v.sum()) if hasattr(v, "sum") else float(v)

def _nt_mens(fc_v2, series_v2, seg, yr, annee_fin):
    """Valeur mensuelle NT (12 mois)."""
    if yr <= annee_fin:
        s = series_v2.get(seg)
        if s is None: return np.zeros(12)
        return s[s.index.year == yr].values
    fc = fc_v2.get(yr, {})
    if seg == "NT":      return fc.get("mensuel_nt",  np.zeros(12))
    if seg == "Transbordement": return fc.get("mensuel_tr", np.zeros(12))
    if seg == "Total":   return fc.get("mensuel_tot", np.zeros(12))
    segs = fc.get("segs_nt", {})
    v = segs.get(seg)
    if v is None: return np.zeros(12)
    return np.array(v) if hasattr(v, "__len__") else np.zeros(12)

def _cnt_total_dyn(fc_cnt, mdl_cnt, yr, cle_cnt, annee_fin):
    """Total conteneurs = NT × ratio[cle_cnt]."""
    if yr <= annee_fin:
        return mdl_cnt.get("ann_total", {}).get(yr, 0)
    fc = fc_cnt.get(yr, {})
    nt_ann = fc.get("annuel_nt", fc.get("annuel", 0))
    if nt_ann == 0: return fc.get("annuel", 0)
    ratio_map = mdl_cnt.get("ratio_tot_nt", {})
    r = ratio_map.get(cle_cnt, ratio_map.get(annee_fin, 1.4398))
    return int(round(nt_ann * r))

# ─────────────────────────────────────────────────────────────────
# 3. EN-TÊTE COMMUNE
# ─────────────────────────────────────────────────────────────────
def _entete_lt(ws, annees_fc, NB_COLS, COL_HYP, titre_doc):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28

    # Titre document
    ws.merge_cells(f"A1:{_col_letter(NB_COLS)}1")
    _c(ws, 1, 1, titre_doc, bg=C_TOTAL, fg="FFFFFF",
       bold=True, size=13, align="center", border=False)
    ws.row_dimensions[1].height = 26

    # En-têtes colonnes
    _c(ws, 2, 1, "N°",      bg=C_TOTAL, fg="FFFFFF", bold=True)
    _c(ws, 2, 2, "RUBRIQUE",bg=C_TOTAL, fg="FFFFFF", bold=True, align="left")
    _c(ws, 2, 3, "RÉALISÉ", bg=C_TOTAL, fg="FFFFFF", bold=True)
    for i, yr in enumerate(annees_fc):
        _c(ws, 2, 4+i, str(yr), bg=C_TOTAL, fg="FFFFFF", bold=True)
    _c(ws, 2, COL_HYP, "HYPOTHÈSE / MÉTHODE",
       bg=C_TOTAL, fg="FFFFFF", bold=True, align="left")
    ws.row_dimensions[2].height = 20
    ws.column_dimensions[_col_letter(3)].width = 14
    for i in range(len(annees_fc)):
        ws.column_dimensions[_col_letter(4+i)].width = 12
    ws.column_dimensions[_col_letter(COL_HYP)].width = 45

def _entete_ct(ws, noms_m, NB_COLS, COL_TOT, COL_HYP, titre_doc, annee_fc):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28

    ws.merge_cells(f"A1:{_col_letter(NB_COLS)}1")
    _c(ws, 1, 1, titre_doc, bg=C_TOTAL, fg="FFFFFF",
       bold=True, size=13, align="center", border=False)
    ws.row_dimensions[1].height = 26

    _c(ws, 2, 1, "N°",      bg=C_TOTAL, fg="FFFFFF", bold=True)
    _c(ws, 2, 2, "RUBRIQUE",bg=C_TOTAL, fg="FFFFFF", bold=True, align="left")
    for m, nom in enumerate(noms_m):
        _c(ws, 2, 3+m, nom, bg=C_TOTAL, fg="FFFFFF", bold=True)
    _c(ws, 2, COL_TOT, str(annee_fc), bg=C_TOTAL, fg="FFFFFF", bold=True)
    _c(ws, 2, COL_HYP, "HYPOTHÈSE / MÉTHODE",
       bg=C_TOTAL, fg="FFFFFF", bold=True, align="left")
    ws.row_dimensions[2].height = 20
    for m in range(12):
        ws.column_dimensions[_col_letter(3+m)].width = 8
    ws.column_dimensions[_col_letter(COL_TOT)].width = 12
    ws.column_dimensions[_col_letter(COL_HYP)].width = 45

# ─────────────────────────────────────────────────────────────────
# 4. SECTION MARCHANDISES LT
# ─────────────────────────────────────────────────────────────────
def _section_march_lt(ws, fc_v2, series_v2, meta, annee_fin,
                      annees_fc, row, COL_HYP, cle_march):
    """Écrit la section marchandises V2 (NT + Transb + Total + axes NT)."""

    parts_ref = meta["parts_nt"].get(cle_march,
                meta["parts_nt"].get(annee_fin, {}))
    wmape = meta["wmape_nt"]

    def hyp_td(seg):
        p = parts_ref.get(seg, 0)
        label = SEGS_NT_LABEL.get(seg, seg)
        return (f"Top-down : NT × {p:.2f}%"
                f" (part {label} en {cle_march})")

    # Définition des lignes par sous-section
    SECTIONS = [
        {
            "titre" : "Trafic global (en Mt)",
            "col_a" : "Trafic global\n(en tonnes)",
            "lignes": [
                ("Total", "Trafic global",
                 C_TOTAL, "FFFFFF", True,
                 f"Total = Non transbordé + Transbordé"),
                ("NT", "Non transbordé (National + Transit)",
                 C_NT, C_GRP1_FG, True,
                 f"SARIMA(1,0,0)(0,1,0)12 · WMAPE={wmape:.1f}% · train 2015-{annee_fin}"),
                ("Transbordement", "Transbordé",
                 C_TR, C_GRP1_FG, True,
                 f"Constante = moyenne 2024-2025 = {meta['transb_moy']:.3f} Mt/an"),
            ]
        },
        {
            "titre" : "Sens du trafic NT (en Mt)",
            "col_a" : "Sens trafic NT\n(en tonnes)",
            "lignes": [
                ("NT", "Non transbordé",
                 C_NT, C_GRP1_FG, True,
                 f"SARIMA(1,0,0)(0,1,0)12 · WMAPE={wmape:.1f}%"),
                ("Import", "Importations",
                 C_BLANC, C_PREV_FG, False, hyp_td("Import")),
                ("Export", "Exportations",
                 C_BLANC, C_PREV_FG, False, hyp_td("Export")),
            ]
        },
        {
            "titre" : "Composante NT (en Mt)",
            "col_a" : "Composante NT\n(en tonnes)",
            "lignes": [
                ("NT", "Non transbordé",
                 C_NT, C_GRP1_FG, True,
                 f"SARIMA(1,0,0)(0,1,0)12 · WMAPE={wmape:.1f}%"),
                ("March. generales", "March. générales",
                 C_BLANC, C_PREV_FG, False,
                 hyp_td("March. generales")),
                ("Prod. petroliers", "Prod. pétroliers",
                 C_BLANC, C_PREV_FG, False,
                 hyp_td("Prod. petroliers")),
                ("Prod. de peche", "Prod. de pêche",
                 C_BLANC, C_PREV_FG, False,
                 hyp_td("Prod. de peche")),
            ]
        },
        {
            "titre" : "Conteneurisation NT (en Mt)",
            "col_a" : "Conteneurisation NT\n(en tonnes)",
            "lignes": [
                ("NT", "Non transbordé",
                 C_NT, C_GRP1_FG, True,
                 f"SARIMA(1,0,0)(0,1,0)12 · WMAPE={wmape:.1f}%"),
                ("Conteneurise", "Conteneurisé",
                 C_BLANC, C_PREV_FG, False,
                 hyp_td("Conteneurise")),
                ("Non conteneurise", "Non conteneurisé",
                 C_BLANC, C_PREV_FG, False,
                 hyp_td("Non conteneurise")),
            ]
        },
    ]

    num = 1
    for section in SECTIONS:
        # Titre section
        ws.merge_cells(f"A{row}:B{row}")
        _c(ws, row, 1, section["titre"],
           bg=C_GRP1, fg=C_GRP1_FG, bold=True, align="left")
        for col in range(3, COL_HYP+1):
            _c(ws, row, col, "", bg=C_GRP1)
        ws.row_dimensions[row].height = 19
        row += 1

        row_start = row
        for seg, label, bg, fg, bold, hyp in section["lignes"]:
            # Réalisé
            v_r = _nt_ann(fc_v2, series_v2, seg, annee_fin, annee_fin)
            _c(ws, row, 1, "", bg=bg)
            _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
            _c(ws, row, 3, round(float(v_r), 3),
               bg=bg, fg=C_REEL_FG if seg in ("Total","NT") else fg,
               bold=bold, num_fmt="#,##0.000")
            # Prévisions
            for i, yr in enumerate(annees_fc):
                v = _nt_ann(fc_v2, series_v2, seg, yr, annee_fin)
                # Recalculer avec cle_march si différente
                if yr > annee_fin and seg not in ("Total","NT","Transbordement"):
                    nt_yr = _nt_ann(fc_v2, series_v2, "NT", yr, annee_fin)
                    p = parts_ref.get(seg, 0)
                    v = round(float(nt_yr) * p / 100, 3)
                _c(ws, row, 4+i, round(float(v), 3),
                   bg=bg, fg=fg, bold=bold, num_fmt="#,##0.000")
            # Hypothèse
            _c(ws, row, COL_HYP, hyp,
               bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
            ws.row_dimensions[row].height = 16
            row += 1

        # Merge colonne A
        if row > row_start + 1:
            ws.merge_cells(f"A{row_start}:A{row-1}")
        c = ws.cell(row_start, 1)
        c.value = section["col_a"]
        c.fill  = PatternFill("solid", fgColor=C_NT)
        c.font  = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True, textRotation=90)

        ws.row_dimensions[row].height = 6
        row += 1

    return row

# ─────────────────────────────────────────────────────────────────
# 5. SECTION MARCHANDISES CT
# ─────────────────────────────────────────────────────────────────
def _section_march_ct(ws, fc_v2, series_v2, meta, annee_fin,
                      annee_fc, row, COL_TOT, COL_HYP, cle_march):
    parts_ref = meta["parts_nt"].get(cle_march,
                meta["parts_nt"].get(annee_fin, {}))
    wmape = meta["wmape_nt"]
    noms_m = meta["noms_mois"]

    def hyp_td(seg):
        p = parts_ref.get(seg, 0)
        label = SEGS_NT_LABEL.get(seg, seg)
        return f"Top-down : NT × {p:.2f}% (part {label} en {cle_march})"

    SECTIONS = [
        {
            "titre" : "Trafic global (en Mt)",
            "col_a" : "Trafic global\n(en tonnes)",
            "lignes": [
                ("Total",         "Trafic global",         C_TOTAL_LINE, C_TOTAL_LINE_FG, True,
                 "Total = NT + Transbordé"),
                ("NT",            "Non transbordé",         C_NT, C_NT_FG, True,
                 f"SARIMA(1,0,0)(0,1,0)12 · WMAPE={wmape:.1f}%"),
                ("Transbordement","Transbordé",             C_TR, C_TR_FG, True,
                 f"Constante = {meta['transb_moy']:.3f} Mt/an"),
            ]
        },
        {
            "titre" : "Sens du trafic NT (en Mt)",
            "col_a" : "Sens NT\n(tonnes)",
            "lignes": [
                ("NT",     "Non transbordé", C_NT, C_NT_FG, True,
                 f"SARIMA · WMAPE={wmape:.1f}%"),
                ("Import", "Importations",   C_BLANC, C_PREV_FG, False,
                 hyp_td("Import")),
                ("Export", "Exportations",   C_BLANC, C_PREV_FG, False,
                 hyp_td("Export")),
            ]
        },
        {
            "titre" : "Composante NT (en Mt)",
            "col_a" : "Composante NT\n(tonnes)",
            "lignes": [
                ("NT",               "Non transbordé",   C_NT, C_NT_FG, True,
                 f"SARIMA · WMAPE={wmape:.1f}%"),
                ("March. generales", "March. générales", C_BLANC,C_PREV_FG, False,
                 hyp_td("March. generales")),
                ("Prod. petroliers", "Prod. pétroliers", C_BLANC,C_PREV_FG, False,
                 hyp_td("Prod. petroliers")),
                ("Prod. de peche",   "Prod. de pêche",   C_BLANC,C_PREV_FG, False,
                 hyp_td("Prod. de peche")),
            ]
        },
        {
            "titre" : "Conteneurisation NT (en Mt)",
            "col_a" : "Conteneur. NT\n(tonnes)",
            "lignes": [
                ("NT",              "Non transbordé",  C_NT, C_NT_FG, True,
                 f"SARIMA · WMAPE={wmape:.1f}%"),
                ("Conteneurise",    "Conteneurisé",    C_BLANC,C_PREV_FG, False,
                 hyp_td("Conteneurise")),
                ("Non conteneurise","Non conteneurisé",C_BLANC,C_PREV_FG, False,
                 hyp_td("Non conteneurise")),
            ]
        },
    ]

    for section in SECTIONS:
        ws.merge_cells(f"A{row}:B{row}")
        _c(ws, row, 1, section["titre"],
           bg=C_GRP1, fg=C_GRP1_FG, bold=True, align="left")
        for col in range(3, COL_HYP+1):
            _c(ws, row, col, "", bg=C_GRP1)
        ws.row_dimensions[row].height = 19
        row += 1

        row_start = row
        for seg, label, bg, fg, bold, hyp in section["lignes"]:
            fc_m = _nt_mens(fc_v2, series_v2, seg, annee_fc, annee_fin)
            # Recalculer avec cle_march si différente
            if seg not in ("Total", "NT", "Transbordement"):
                nt_m = _nt_mens(fc_v2, series_v2, "NT", annee_fc, annee_fin)
                p = parts_ref.get(seg, 0)
                fc_m = np.array([v * p / 100 for v in nt_m])
            fc_m = [round(float(v), 3) for v in fc_m]

            _c(ws, row, 1, "", bg=bg)
            _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
            for m in range(12):
                _c(ws, row, 3+m, fc_m[m], bg=bg, fg=fg, bold=bold,
                   num_fmt="#,##0.000")
            _c(ws, row, COL_TOT, round(sum(fc_m), 3),
               bg=bg, fg=fg, bold=True, num_fmt="#,##0.000")
            _c(ws, row, COL_HYP, hyp,
               bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
            ws.row_dimensions[row].height = 16
            row += 1

        if row > row_start + 1:
            ws.merge_cells(f"A{row_start}:A{row-1}")
        c = ws.cell(row_start, 1)
        c.value = section["col_a"]
        c.fill  = PatternFill("solid", fgColor=C_NT)
        c.font  = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
        c.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True, textRotation=90)
        ws.row_dimensions[row].height = 6
        row += 1

    return row

# ─────────────────────────────────────────────────────────────────
# 6. SECTION ESCALES (reprise V1)
# ─────────────────────────────────────────────────────────────────
def _section_esc_lt(ws, fc_esc, mdl_esc, annee_fin,
                    annees_fc, row, COL_HYP, cle_esc):
    if not fc_esc or not mdl_esc: return row
    ann_hist = mdl_esc["ann_total_hist"]
    yr_last  = mdl_esc["annee_fin"]
    wmape    = mdl_esc["wmape"]
    parts_yr = mdl_esc["parts_terminaux"]
    _cle_r   = cle_esc if cle_esc else yr_last

    def _hyp(g):
        p = parts_yr.get(_cle_r, parts_yr[yr_last]).get(g, 0)
        return (f"Top-down : Nb total escales × {p:.2f}%"
                f" (part {g} en {_cle_r})")

    SEGS_ESC = list(parts_yr.get(yr_last, {}).keys())

    def esc_v(cle, yr):
        if yr == annee_fin:
            if cle == "TOTAL": return ann_hist.get(yr, 0)
            return int(fc_esc[("historique", yr)]["segments"]
                       .get(cle, np.zeros(12)).sum())
        if cle == "TOTAL": return fc_esc[yr]["annuel"]
        total_yr = fc_esc[yr]["annuel"]
        p = parts_yr.get(_cle_r, parts_yr[yr_last]).get(cle, 0)
        return int(round(total_yr * p / 100))

    LIGNES = [
        ("TOTAL",        "Nombre total d'escales",           C_TOTAL, "FFFFFF", True,
         f"Holt amorti (double lissage) · WMAPE={wmape:.1f}%"),

        ("TC1",          "TERMINAL A CONTENEUR (TC 1)",       C_BLANC, C_PREV_FG, True,
         _hyp("TC1")),
        ("TC2",          "TERMINAL A CONTENEUR (TC 2)",       C_BLANC, C_PREV_FG, True,
         _hyp("TC2")),
        ("Céréalier",    "TERMINAL CÉRÉALIER",                C_BLANC, C_PREV_FG, True,
         _hyp("Céréalier")),
        ("Fruitier",     "TERMINAL FRUITIER",                 C_BLANC, C_PREV_FG, True,
         _hyp("Fruitier")),
        ("Minéralier",   "TERMINAL MINÉRALIER",               C_BLANC, C_PREV_FG, True,
         _hyp("Minéralier")),
        ("Pétrolier",    "TERMINAL PÉTROLIER",                C_BLANC, C_PREV_FG, True,
         _hyp("Pétrolier")),
        ("Roulier",      "TERMINAL ROULIER",                  C_BLANC, C_PREV_FG, True,
         _hyp("Roulier")),
        ("Quai Nord",    "Quai NORD (Q1-Q5)",                 C_BLANC, C_PREV_FG, True,
         _hyp("Quai Nord")),
        ("Quai Ouest",   "Quai OUEST (Q6-Q10)",               C_BLANC, C_PREV_FG, True,
         _hyp("Quai Ouest")),
        ("Autres zones", "AUTRES ZONES",                      C_BLANC, C_PREV_FG, True,
         _hyp("Autres zones")),
    ]

    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "Trafic global de navires (en nombre)",
       bg=C_ESC, fg=C_ESC_FG, bold=True, align="left")
    for col in range(3, COL_HYP+1):
        _c(ws, row, col, "", bg=C_GRP2)
    ws.row_dimensions[row].height = 19
    row += 1

    row_start = row
    for cle, label, bg, fg, bold, hyp in LIGNES:
        v_r = esc_v(cle, annee_fin)
        _c(ws, row, 1, "", bg=bg)
        _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
        _c(ws, row, 3, int(v_r), bg=bg,
           fg=C_REEL_FG if cle in ("TOTAL",) else ("FFFFFF" if fg not in ("000000","1B2631") else "CC0000"),
           bold=bold, num_fmt="#,##0")
        for i, yr in enumerate(annees_fc):
            _c(ws, row, 4+i, esc_v(cle, yr),
               bg=bg, fg=fg, bold=bold, num_fmt="#,##0")
        _c(ws, row, COL_HYP, hyp,
           bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.merge_cells(f"A{row_start}:A{row-1}")
    c = ws.cell(row_start, 1)
    c.value = "Trafic global\nde navires"
    c.fill  = PatternFill("solid", fgColor=C_ESC)
    c.font  = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True, textRotation=90)
    ws.row_dimensions[row].height = 6
    return row + 1

def _section_esc_ct(ws, fc_esc, mdl_esc, annee_fin, annee_fc,
                    row, COL_TOT, COL_HYP, cle_esc):
    if not fc_esc or not mdl_esc: return row
    yr_last  = mdl_esc["annee_fin"]
    wmape    = mdl_esc["wmape"]
    parts_yr = mdl_esc["parts_terminaux"]
    _cle_r   = cle_esc if cle_esc else yr_last

    def _hyp(g):
        p = parts_yr.get(_cle_r, parts_yr[yr_last]).get(g, 0)
        return f"Top-down : Nb total escales × {p:.2f}% (part {g} en {_cle_r})"

    def em(cle):
        total_m = np.array(fc_esc[annee_fc]["mensuel"], dtype=float)
        if cle == "TOTAL": return total_m
        p = parts_yr.get(_cle_r, parts_yr[yr_last]).get(cle, 0)
        r = np.round(total_m * p / 100).astype(int).astype(float)
        diff = total_m.sum() * p / 100 - r.sum()
        if abs(diff) >= 0.5: r[int(np.argmax(total_m))] += round(diff)
        return r

    LIGNES = [
        ("TOTAL",        "Nombre total d'escales",           C_TOTAL, "FFFFFF", True,
         f"Holt amorti · WMAPE={wmape:.1f}%"),

        ("TC1",          "TERMINAL A CONTENEUR (TC 1)",       C_BLANC, C_PREV_FG, True,
         _hyp("TC1")),
        ("TC2",          "TERMINAL A CONTENEUR (TC 2)",       C_BLANC, C_PREV_FG, True,
         _hyp("TC2")),
        ("Céréalier",    "TERMINAL CÉRÉALIER",                C_BLANC, C_PREV_FG, True,
         _hyp("Céréalier")),
        ("Fruitier",     "TERMINAL FRUITIER",                 C_BLANC, C_PREV_FG, True,
         _hyp("Fruitier")),
        ("Minéralier",   "TERMINAL MINÉRALIER",               C_BLANC, C_PREV_FG, True,
         _hyp("Minéralier")),
        ("Pétrolier",    "TERMINAL PÉTROLIER",                C_BLANC, C_PREV_FG, True,
         _hyp("Pétrolier")),
        ("Roulier",      "TERMINAL ROULIER",                  C_BLANC, C_PREV_FG, True,
         _hyp("Roulier")),
        ("Quai Nord",    "Quai NORD (Q1-Q5)",                 C_BLANC, C_PREV_FG, True,
         _hyp("Quai Nord")),
        ("Quai Ouest",   "Quai OUEST (Q6-Q10)",               C_BLANC, C_PREV_FG, True,
         _hyp("Quai Ouest")),
        ("Autres zones", "AUTRES ZONES",                      C_BLANC, C_PREV_FG, True,
         _hyp("Autres zones")),
    ]

    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "Trafic global de navires (en nombre)",
       bg=C_ESC, fg=C_ESC_FG, bold=True, align="left")
    for col in range(3, COL_HYP+1):
        _c(ws, row, col, "", bg=C_GRP2)
    ws.row_dimensions[row].height = 19
    row += 1

    row_start = row
    for cle, label, bg, fg, bold, hyp in LIGNES:
        fc_m = em(cle)
        _c(ws, row, 1, "", bg=bg)
        _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
        for m in range(12):
            _c(ws, row, 3+m, int(fc_m[m]), bg=bg, fg=fg, bold=bold,
               num_fmt="#,##0")
        _c(ws, row, COL_TOT, int(fc_m.sum()), bg=bg, fg=fg, bold=True,
           num_fmt="#,##0")
        _c(ws, row, COL_HYP, hyp,
           bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.merge_cells(f"A{row_start}:A{row-1}")
    c = ws.cell(row_start, 1)
    c.value = "Trafic global\nde navires"
    c.fill  = PatternFill("solid", fgColor=C_ESC)
    c.font  = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True, textRotation=90)
    ws.row_dimensions[row].height = 6
    return row + 1

# ─────────────────────────────────────────────────────────────────
# 7. SECTION CONTENEURS V2 LT
# ─────────────────────────────────────────────────────────────────
LIGNES_CNT = [
    ("TOTAL",     "Trafic total conteneurs",             C_TOTAL_LINE, C_TOTAL_LINE_FG, True),
    ("NT",        "Non transbordé",                      C_CNT_NT, C_NT_FG, True),
    ("Transb",    "Transbordés (total)",                 C_CNT_TR, C_CNT_TR_FG, True),
    ("TransbTC2", "Transbordés TC2",                     C_BLANC,C_PREV_FG,False),
    ("TransbHab", "Transbordés habituel",                C_BLANC,C_PREV_FG,False),
    ("TC1",       "TERMINAL A CONTENEUR (TC 1)",         C_BLANC,C_PREV_FG,False),
    ("TC2",       "TERMINAL A CONTENEUR (TC 2)",         C_BLANC,C_PREV_FG,False),
    ("Fruitier",  "TERMINAL FRUITIER",                   C_BLANC,C_PREV_FG,False),
    ("Roulier",   "TERMINAL ROULIER",                    C_BLANC,C_PREV_FG,False),
    ("Autres",    "AUTRES ZONES",                        C_BLANC,C_PREV_FG,False),
]

def _cnt_v2_val(cle, yr, fc_cnt, mdl_cnt, annee_fin, cle_cnt):
    """Valeur annuelle conteneurs V2."""
    import numpy as np

    # ── Réalisé (historique) ──────────────────────────────────────
    if yr == annee_fin:
        fh = fc_cnt.get(("historique", yr), {})
        segs_t = fh.get("segments_term", {})
        segs_d = fh.get("segments_dest", {})
        tot = fh.get("annuel", 0)
        map_t = {"TC1":"TC1","TC2":"TC2","Fruitier":"Fruitier",
                 "Roulier":"Roulier","Autres":"Autres zones"}
        if cle == "TOTAL":     return tot
        if cle == "NT":
            return int(np.array(segs_d.get("Non transb.", [0])).sum()
                       if "Non transb." in segs_d else int(mdl_cnt["y_nt"][-1]))
        if cle == "Transb":
            return int(np.array(segs_d.get("Transbordé", [0])).sum()
                       if "Transbordé" in segs_d else 0)
        if cle == "TransbTC2":
            return int(np.array(segs_d.get("Transb. TC2", [0])).sum()
                       if "Transb. TC2" in segs_d else mdl_cnt["transb_tc2_2025"])
        if cle == "TransbHab":
            return int(np.array(segs_d.get("Transb. habituel", [0])).sum()
                       if "Transb. habituel" in segs_d else 0)
        t_key = map_t.get(cle, cle)
        if t_key in segs_t:
            return int(np.array(segs_t[t_key]).sum())
        return 0

    # ── Prévisions ───────────────────────────────────────────────
    fc = fc_cnt.get(yr, {})
    tot     = _cnt_total_dyn(fc_cnt, mdl_cnt, yr, cle_cnt, annee_fin)
    nt_ann  = fc.get("annuel_nt", 0)
    tc2_ann = fc.get("transb_tc2_ann", mdl_cnt.get("transb_tc2_2025", 0))
    parts_t = mdl_cnt.get("parts_term", {}).get(cle_cnt,
              mdl_cnt.get("parts_term", {}).get(annee_fin, {}))
    map_t = {"TC1":"TC1","TC2":"TC2","Fruitier":"Fruitier",
             "Roulier":"Roulier","Autres":"Autres zones"}

    if cle == "TOTAL":     return tot
    if cle == "NT":        return nt_ann
    if cle == "Transb":    return tot - nt_ann
    if cle == "TransbTC2": return tc2_ann
    if cle == "TransbHab": return tot - nt_ann - tc2_ann
    return int(round(tot * parts_t.get(map_t.get(cle, cle), 0) / 100))


def _section_cnt_lt_v2(ws, fc_cnt, mdl_cnt, annees_fc,
                       annee_fin, row, COL_HYP, cle_cnt):
    wmape_nt = mdl_cnt.get("wmape_nt", 0)
    parts_t  = mdl_cnt.get("parts_term", {}).get(cle_cnt,
               mdl_cnt.get("parts_term", {}).get(annee_fin, {}))
    parts_d  = mdl_cnt.get("parts_dest", {}).get(cle_cnt,
               mdl_cnt.get("parts_dest", {}).get(annee_fin, {}))
    ratio_map= mdl_cnt.get("ratio_tot_nt", {})
    r_val    = ratio_map.get(cle_cnt, ratio_map.get(annee_fin, 1.4398))
    tc2_val  = mdl_cnt.get("transb_tc2_2025", 0)

    def _hyp(cle):
        if cle == "TOTAL":
            return f"NT × ratio {r_val:.4f} (Total/NT en {cle_cnt})"
        if cle == "NT":
            return f"Holt amorti · WMAPE={wmape_nt:.1f}% · train 2015-{annee_fin}"
        if cle == "Transb":
            return "Somme : Transbordés TC2 + Transbordés habituel"
        if cle == "TransbTC2":
            return f"Constant = réalisé {annee_fin} : {tc2_val:,} EVP/an"
        if cle == "TransbHab":
            return "Résiduel : Total − (Non transbordé + Transbordé TC2)"
        map_t = {"TC1":"TC1","TC2":"TC2","Fruitier":"Fruitier",
                 "Roulier":"Roulier","Autres":"Autres zones"}
        p = parts_t.get(map_t.get(cle,cle), 0)
        return f"Top-down : Total × {p:.2f}% (part {cle} en {cle_cnt})"

    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "Trafic de conteneurs (en EVP)",
       bg=C_CNT_TOT, fg=C_CNT_FG, bold=True, align="left")
    for col in range(3, COL_HYP+1):
        _c(ws, row, col, "", bg=C_TR)
    ws.row_dimensions[row].height = 19
    row += 1

    row_start = row
    for cle, label, bg, fg, bold in LIGNES_CNT:
        v_r = _cnt_v2_val(cle, annee_fin, fc_cnt, mdl_cnt, annee_fin, cle_cnt)
        _c(ws, row, 1, "", bg=bg)
        _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
        _c(ws, row, 3, int(v_r), bg=bg,
           fg=C_REEL_FG if cle in ("TOTAL",) else ("FFFFFF" if fg not in ("000000","1B2631") else "CC0000"),
           bold=bold, num_fmt="#,##0")
        for i, yr in enumerate(annees_fc):
            v = _cnt_v2_val(cle, yr, fc_cnt, mdl_cnt, annee_fin, cle_cnt)
            _c(ws, row, 4+i, int(v), bg=bg, fg=fg, bold=bold, num_fmt="#,##0")
        _c(ws, row, COL_HYP, _hyp(cle),
           bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.merge_cells(f"A{row_start}:A{row-1}")
    c = ws.cell(row_start, 1)
    c.value = "Trafic de\nconteneurs\n(EVP)"
    c.fill  = PatternFill("solid", fgColor=C_CNT_TOT)
    c.font  = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True, textRotation=90)
    ws.row_dimensions[row].height = 6
    return row + 1

def _section_cnt_ct_v2(ws, fc_cnt, mdl_cnt, annee_fin, annee_fc,
                       row, COL_TOT, COL_HYP, cle_cnt):
    wmape_nt = mdl_cnt.get("wmape_nt", 0)
    parts_t  = mdl_cnt.get("parts_term", {}).get(cle_cnt,
               mdl_cnt.get("parts_term", {}).get(annee_fin, {}))
    ratio_map= mdl_cnt.get("ratio_tot_nt", {})
    r_val    = ratio_map.get(cle_cnt, ratio_map.get(annee_fin, 1.4398))
    tc2_val  = mdl_cnt.get("transb_tc2_2025", 0)

    fc = fc_cnt.get(annee_fc, {})
    mens_tot = np.array(fc.get("mensuel", np.zeros(12)), dtype=float)
    # Recalibrer selon ratio de la clé
    nt_ann = fc.get("annuel_nt", 0)
    if nt_ann > 0 and mens_tot.sum() > 0:
        tot_dyn = int(round(nt_ann * r_val))
        factor  = tot_dyn / mens_tot.sum()
        mens_tot = np.round(mens_tot * factor).astype(int).astype(float)

    def get_mens_cnt(cle):
        if cle == "TOTAL":  return mens_tot
        if cle == "NT":     return np.array(fc.get("mensuel",np.zeros(12)),dtype=float) * (nt_ann/mens_tot.sum() if mens_tot.sum()>0 else 1)
        tc2_m = np.round(tc2_val/12 * np.ones(12)).astype(int).astype(float)
        if cle == "TransbTC2": return tc2_m
        # NT mensuel recalibré selon ratio
        total_base = fc.get("mensuel", np.zeros(12))
        tot_sum    = mens_tot.sum()
        nt_m_cal   = np.round(mens_tot * (nt_ann / tot_sum)).astype(int).astype(float)                      if tot_sum > 0 and nt_ann > 0 else np.zeros(12)
        if cle == "Transb":    return mens_tot - nt_m_cal
        if cle == "TransbHab": return mens_tot - nt_m_cal - tc2_m
        map_t = {"TC1":"TC1","TC2":"TC2","Fruitier":"Fruitier",
                 "Roulier":"Roulier","Autres":"Autres zones"}
        segs = fc.get("segments_term", {})
        if map_t.get(cle,cle) in segs:
            s_m = np.array(segs[map_t.get(cle,cle)], dtype=float)
            if mens_tot.sum() > 0:
                s_m = np.round(s_m * (mens_tot.sum()/fc.get("annuel",mens_tot.sum()))).astype(int).astype(float)
            return s_m
        p = parts_t.get(map_t.get(cle,cle), 0)
        return np.round(mens_tot * p / 100).astype(int).astype(float)

    def _hyp(cle):
        if cle == "TOTAL":     return f"NT × ratio {r_val:.4f} (Total/NT en {cle_cnt})"
        if cle == "NT":        return f"Holt amorti · WMAPE={wmape_nt:.1f}%"
        if cle == "Transb":    return "Somme TC2 + Habituel"
        if cle == "TransbTC2": return f"Constant : {tc2_val:,} EVP/an"
        if cle == "TransbHab": return "Résiduel : Total − (NT + TC2)"
        map_t = {"TC1":"TC1","TC2":"TC2","Fruitier":"Fruitier",
                 "Roulier":"Roulier","Autres":"Autres zones"}
        p = parts_t.get(map_t.get(cle,cle), 0)
        return f"Top-down : Total × {p:.2f}% (part {cle} en {cle_cnt})"

    ws.merge_cells(f"A{row}:B{row}")
    _c(ws, row, 1, "Trafic de conteneurs (en EVP)",
       bg=C_CNT_TOT, fg=C_CNT_FG, bold=True, align="left")
    for col in range(3, COL_HYP+1):
        _c(ws, row, col, "", bg=C_TR)
    ws.row_dimensions[row].height = 19
    row += 1

    row_start = row
    for cle, label, bg, fg, bold in LIGNES_CNT:
        fc_m = get_mens_cnt(cle)
        _c(ws, row, 1, "", bg=bg)
        reel_fg = ("FF0000" if cle=="TOTAL" else
                   "FFFF99" if bg not in (C_BLANC,"FFFFFF") else "CC0000")
        _c(ws, row, 2, label, bg=bg, fg=fg, bold=bold, align="left")
        for m in range(12):
            _c(ws, row, 3+m, int(fc_m[m]), bg=bg, fg=fg, bold=bold,
               num_fmt="#,##0")
        _c(ws, row, COL_TOT, int(fc_m.sum()), bg=bg, fg=fg, bold=True,
           num_fmt="#,##0")
        _c(ws, row, COL_HYP, _hyp(cle),
           bg=C_BLANC, fg=C_PREV_FG, italic=True, size=9, align="left")
        ws.row_dimensions[row].height = 16
        row += 1

    ws.merge_cells(f"A{row_start}:A{row-1}")
    c = ws.cell(row_start, 1)
    c.value = "Trafic de\nconteneurs\n(EVP)"
    c.fill  = PatternFill("solid", fgColor=C_CNT_TOT)
    c.font  = Font(color="FFFFFF", bold=True, size=9, name="Calibri")
    c.alignment = Alignment(horizontal="center", vertical="center",
                             wrap_text=True, textRotation=90)
    ws.row_dimensions[row].height = 6
    return row + 1

# ─────────────────────────────────────────────────────────────────
# 8. FONCTIONS PUBLIQUES
# ─────────────────────────────────────────────────────────────────
def generate_xlsx_lt_v2(forecasts_v2, meta_v2, series_v2,
                        fc_esc, mdl_esc, fc_cnt, mdl_cnt,
                        annee_max_data, annee_min_fc, horizon,
                        cle_march=None, cle_esc=None, cle_cnt=None):
    """Génère le tableau long terme V2."""
    cle_march = cle_march or annee_max_data
    cle_esc   = cle_esc   or annee_max_data
    cle_cnt   = cle_cnt   or annee_max_data

    annees_fc = list(range(annee_min_fc, annee_min_fc + horizon))
    NB_COLS   = 3 + horizon + 1   # A + B + réalisé + prévisions + hypothèse
    COL_HYP   = NB_COLS

    wb = Workbook()
    ws = wb.active
    ws.title = f"Prévisions LT {annee_min_fc}-{annee_min_fc+horizon-1}"

    titre = (f"PORT AUTONOME D'ABIDJAN — Prévisions long terme "
             f"{annee_min_fc}–{annee_min_fc+horizon-1} (App V2 · NT + Transbordé)")
    _entete_lt(ws, annees_fc, NB_COLS, COL_HYP, titre)

    row = 3
    # Marchandises
    row = _section_march_lt(ws, forecasts_v2, series_v2, meta_v2,
                             annee_max_data, annees_fc, row, COL_HYP, cle_march)
    # Escales
    row = _section_esc_lt(ws, fc_esc, mdl_esc, annee_max_data,
                           annees_fc, row, COL_HYP, cle_esc)
    # Conteneurs V2
    row = _section_cnt_lt_v2(ws, fc_cnt, mdl_cnt, annees_fc,
                              annee_max_data, row, COL_HYP, cle_cnt)

    ws.freeze_panes = "C3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_xlsx_ct_v2(forecasts_v2, meta_v2, series_v2,
                        fc_esc, mdl_esc, fc_cnt, mdl_cnt,
                        annee_max_data, annee_fc,
                        cle_march=None, cle_esc=None, cle_cnt=None):
    """Génère le tableau court terme V2."""
    cle_march = cle_march or annee_max_data
    cle_esc   = cle_esc   or annee_max_data
    cle_cnt   = cle_cnt   or annee_max_data

    noms_m  = meta_v2["noms_mois"]
    NB_COLS = 3 + 12 + 1   # A + B + 12 mois + total + hypothèse
    COL_TOT = 3 + 12
    COL_HYP = NB_COLS

    wb = Workbook()
    ws = wb.active
    ws.title = f"Prévisions CT {annee_fc}"

    titre = (f"PORT AUTONOME D'ABIDJAN — Prévisions court terme "
             f"{annee_fc} (App V2 · NT + Transbordé)")
    _entete_ct(ws, noms_m, NB_COLS, COL_TOT, COL_HYP, titre, annee_fc)

    row = 3
    row = _section_march_ct(ws, forecasts_v2, series_v2, meta_v2,
                             annee_max_data, annee_fc, row,
                             COL_TOT, COL_HYP, cle_march)
    row = _section_esc_ct(ws, fc_esc, mdl_esc, annee_max_data, annee_fc,
                           row, COL_TOT, COL_HYP, cle_esc)
    row = _section_cnt_ct_v2(ws, fc_cnt, mdl_cnt, annee_max_data, annee_fc,
                              row, COL_TOT, COL_HYP, cle_cnt)

    ws.freeze_panes = "C3"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
